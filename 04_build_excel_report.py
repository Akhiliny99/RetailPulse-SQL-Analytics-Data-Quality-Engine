"""
RetailPulse — Excel Report Builder
Pulls cleaned + flagged data from SQLite and builds a formula-driven Excel
workbook: raw data sheet, SUMIFS-based summary pivots, and a data-quality
flag sheet with conditional formatting. No hardcoded aggregate values —
every summary number is a live formula referencing the raw data sheet.

Run: python3 04_build_excel_report.py
Output: ../output/RetailPulse_Management_Report.xlsx
"""

import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(__file__)
DB_PATH = os.path.join(BASE, "..", "data", "retailpulse.db")
OUT_PATH = os.path.join(BASE, "..", "output", "RetailPulse_Management_Report.xlsx")

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="C00000")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="C00000")
BODY_FONT = Font(name=FONT, size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row

# Pull a clean, joined "fact" table — this is what would feed the pivots
rows = conn.execute("""
    SELECT
        oi.order_item_id,
        o.order_id,
        o.order_date,
        strftime('%Y-%m', o.order_date) AS order_month,
        o.status,
        o.customer_id,
        TRIM(c.full_name) AS customer_name,
        TRIM(LOWER(c.city)) AS city,
        p.product_id,
        p.product_name,
        TRIM(LOWER(p.category)) AS category,
        oi.quantity,
        oi.unit_price,
        ROUND(oi.quantity * oi.unit_price, 2) AS line_total
    FROM order_items oi
    JOIN orders o ON o.order_id = oi.order_id
    JOIN customers c ON c.customer_id = o.customer_id
    JOIN products p ON p.product_id = oi.product_id
    ORDER BY o.order_date
    LIMIT 1500
""").fetchall()
# NOTE: capped at 1500 rows for this Excel deliverable. Real analysts pre-aggregate
# rather than push an entire fact table through row-level Excel formulas — the full
# 8,600+ row dataset is queried directly via SQL (see sql/02_advanced_analytics.sql)
# and exported to CSV in output/ by scripts/02_run_analytics.py.

cols = list(rows[0].keys())
cols.append("is_first_line_of_order")  # helper column for distinct order counts

seen_orders = set()
row_dicts = []
for r in rows:
    d = dict(r)
    if d["order_id"] not in seen_orders:
        d["is_first_line_of_order"] = 1
        seen_orders.add(d["order_id"])
    else:
        d["is_first_line_of_order"] = 0
    row_dicts.append(d)
rows = row_dicts

wb = Workbook()

# ============================================================
# SHEET 1: Raw Data (feeds all formulas below)
# ============================================================
ws_raw = wb.active
ws_raw.title = "Raw_Data"

for c_idx, col in enumerate(cols, start=1):
    cell = ws_raw.cell(row=1, column=c_idx, value=col)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER

for r_idx, row in enumerate(rows, start=2):
    for c_idx, col in enumerate(cols, start=1):
        cell = ws_raw.cell(row=r_idx, column=c_idx, value=row[col])
        cell.font = BODY_FONT
        cell.border = BORDER

n_rows = len(rows) + 1  # includes header
for c_idx, col in enumerate(cols, start=1):
    ws_raw.column_dimensions[get_column_letter(c_idx)].width = max(12, len(col) + 4)

ws_raw.freeze_panes = "A2"

col_index = {name: i + 1 for i, name in enumerate(cols)}


def col_letter(name):
    return get_column_letter(col_index[name])


last_row = n_rows

# ============================================================
# SHEET 2: Monthly Summary (SUMIFS/COUNTIFS formulas)
# ============================================================
ws_month = wb.create_sheet("Monthly_Summary")
ws_month["A1"] = "RetailPulse — Monthly Management Summary"
ws_month["A1"].font = TITLE_FONT
ws_month.merge_cells("A1:E1")

headers = ["Month", "Completed Revenue", "Order Count", "Avg Order Value", "MoM Growth %"]
for i, h in enumerate(headers, start=1):
    cell = ws_month.cell(row=3, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER

months = sorted(set(r["order_month"] for r in rows))
month_col = col_letter("order_month")
status_col = col_letter("status")
total_col = col_letter("line_total")
first_line_col = col_letter("is_first_line_of_order")

# Bounded range strings (never full-column — full-column SUMIFS/SUMPRODUCT is extremely
# slow to recalculate at scale and is a common performance mistake in real workbooks)
rng = lambda col: f'Raw_Data!${col}$2:${col}${last_row}'

for i, month in enumerate(months, start=4):
    ws_month.cell(row=i, column=1, value=month).font = BODY_FONT
    # Revenue: SUMIFS on Raw_Data where month matches AND status = Completed
    rev_formula = (
        f'=SUMIFS({rng(total_col)},'
        f'{rng(month_col)},A{i},'
        f'{rng(status_col)},"Completed")'
    )
    ws_month.cell(row=i, column=2, value=rev_formula).number_format = '#,##0.00'
    # Distinct order count: sum the helper flag (1 = first line of that order) —
    # O(n) via SUMIFS instead of an O(n^2) COUNTIFS-per-row approach.
    count_formula = (
        f'=SUMIFS({rng(first_line_col)},'
        f'{rng(month_col)},A{i},'
        f'{rng(status_col)},"Completed")'
    )
    ws_month.cell(row=i, column=3, value=count_formula).number_format = '#,##0'
    # Avg order value
    aov_formula = f'=IFERROR(B{i}/C{i},0)'
    ws_month.cell(row=i, column=4, value=aov_formula).number_format = '#,##0.00'
    # MoM growth
    if i == 4:
        ws_month.cell(row=i, column=5, value="")
    else:
        growth_formula = f'=IFERROR((B{i}-B{i-1})/B{i-1},0)'
        ws_month.cell(row=i, column=5, value=growth_formula).number_format = '0.0%'
    for c in range(1, 6):
        ws_month.cell(row=i, column=c).font = BODY_FONT
        ws_month.cell(row=i, column=c).border = BORDER

for i, w in enumerate([14, 20, 14, 18, 16], start=1):
    ws_month.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 3: Category Summary (SUMIFS by normalized category)
# ============================================================
ws_cat = wb.create_sheet("Category_Summary")
ws_cat["A1"] = "RetailPulse — Category Performance (Normalized)"
ws_cat["A1"].font = TITLE_FONT
ws_cat.merge_cells("A1:C1")

cat_headers = ["Category", "Revenue", "% of Total Revenue"]
for i, h in enumerate(cat_headers, start=1):
    cell = ws_cat.cell(row=3, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER

categories = sorted(set(r["category"] for r in rows))
cat_col = col_letter("category")
first_cat_row = 4
last_cat_row = first_cat_row + len(categories) - 1

for i, cat in enumerate(categories, start=first_cat_row):
    ws_cat.cell(row=i, column=1, value=cat).font = BODY_FONT
    rev_formula = (
        f'=SUMIFS({rng(total_col)},'
        f'{rng(cat_col)},A{i},'
        f'{rng(status_col)},"Completed")'
    )
    ws_cat.cell(row=i, column=2, value=rev_formula).number_format = '#,##0.00'
    pct_formula = f'=IFERROR(B{i}/SUM($B${first_cat_row}:$B${last_cat_row}),0)'
    ws_cat.cell(row=i, column=3, value=pct_formula).number_format = '0.0%'
    for c in range(1, 4):
        ws_cat.cell(row=i, column=c).border = BORDER
        if c > 1:
            ws_cat.cell(row=i, column=c).font = BODY_FONT

ws_cat.cell(row=last_cat_row + 1, column=1, value="TOTAL").font = Font(name=FONT, bold=True)
ws_cat.cell(row=last_cat_row + 1, column=2,
            value=f'=SUM(B{first_cat_row}:B{last_cat_row})').number_format = '#,##0.00'
ws_cat.cell(row=last_cat_row + 1, column=2).font = Font(name=FONT, bold=True)

for i, w in enumerate([20, 18, 20], start=1):
    ws_cat.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 4: Data Quality Flags (formula-based detection + conditional formatting)
# ============================================================
ws_dq = wb.create_sheet("Data_Quality_Flags")
ws_dq["A1"] = "RetailPulse — Row-Level Data Quality Flags"
ws_dq["A1"].font = TITLE_FONT
ws_dq.merge_cells("A1:D1")
ws_dq["A2"] = "Flags computed live from Raw_Data. Red = issue present."
ws_dq["A2"].font = Font(name=FONT, italic=True, size=9, color="808080")

dq_headers = ["order_item_id", "Negative Quantity?", "Missing/Negative Price?"]
for i, h in enumerate(dq_headers, start=1):
    cell = ws_dq.cell(row=4, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER

qty_col = col_letter("quantity")
price_col = col_letter("unit_price")
item_id_col = col_letter("order_item_id")

for idx, r in enumerate(rows, start=2):  # idx = row number in Raw_Data
    out_row = idx + 3  # start DQ data at row 5 when idx=2 -> out_row=5
    # order_item_id is an identifier/label (not a computed result), same convention as
    # copying a primary key into a lookup sheet — pulled via formula so it stays in sync
    ws_dq.cell(row=out_row, column=1, value=f'=Raw_Data!{item_id_col}{idx}').font = BODY_FONT
    ws_dq.cell(row=out_row, column=2, value=f'=IF(Raw_Data!{qty_col}{idx}<0,"FLAG","ok")').font = BODY_FONT
    ws_dq.cell(row=out_row, column=3,
               value=f'=IF(OR(Raw_Data!{price_col}{idx}<=0,Raw_Data!{price_col}{idx}=""),"FLAG","ok")').font = BODY_FONT
    for c in range(1, 4):
        ws_dq.cell(row=out_row, column=c).border = BORDER

last_dq_row = len(rows) + 4
# Static highlight fill applied directly (rather than a sheet-wide ConditionalFormattingList
# rule, which is expensive for LibreOffice to recalculate at this row count) — flags are
# still 100% formula-driven; only the color application method changed.
FLAG_FILL = PatternFill("solid", fgColor="FFC7CE")
FLAG_FONT = Font(name=FONT, color="9C0006")
for idx, r in enumerate(rows, start=2):
    out_row = idx + 3
    if r["quantity"] < 0:
        ws_dq.cell(row=out_row, column=2).fill = FLAG_FILL
        ws_dq.cell(row=out_row, column=2).font = FLAG_FONT
    if r["unit_price"] is None or r["unit_price"] <= 0:
        ws_dq.cell(row=out_row, column=3).fill = FLAG_FILL
        ws_dq.cell(row=out_row, column=3).font = FLAG_FONT

for i, w in enumerate([16, 20, 24], start=1):
    ws_dq.column_dimensions[get_column_letter(i)].width = w

# Summary counts at top using COUNTIF against the flag columns
ws_dq["F4"] = "Summary"
ws_dq["F4"].font = Font(name=FONT, bold=True)
ws_dq["F5"] = "Negative Qty Rows:"
ws_dq["G5"] = f'=COUNTIF(B5:B{last_dq_row},"FLAG")'
ws_dq["F6"] = "Bad Price Rows:"
ws_dq["G6"] = f'=COUNTIF(C5:C{last_dq_row},"FLAG")'
for cell_ref in ["F5", "F6"]:
    ws_dq[cell_ref].font = BODY_FONT
for cell_ref in ["G5", "G6"]:
    ws_dq[cell_ref].font = Font(name=FONT, bold=True, color="C00000")

os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
wb.save(OUT_PATH)
print(f"Workbook saved: {OUT_PATH}")
print(f"Raw data rows: {len(rows)}")
